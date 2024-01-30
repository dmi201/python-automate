import pandas as pd
from openpyxl import Workbook, load_workbook
import xlwings as xw
# import os

# cwd = os.getcwd()
# print(cwd)


wb_template = load_workbook(r'C:\Users\Mihai\Desktop\MyFolder\Sablon.xlsx')
ws_t_gd = wb_template['DateGenerale']
ws_t_a = wb_template['Antene']
ws_t_at = wb_template['Tabel_Antene']
ws_t_gard = wb_template['Gard']
ws_t_fund_t30tri = wb_template['Fundatii_T30tri']

wb_template_f = load_workbook(
    r'C:\Users\Mihai\Desktop\MyFolder\FormuleSablon.xlsx')
ws_t_f_gd = wb_template_f['DateGenerale']
ws_t_f_a = wb_template_f['Antene']
ws_t_f_etc_t30tri = wb_template_f['ExcelToCad_T30tri']
ws_t_f_etc_t20 = wb_template_f['ExcelToCad_T20']
ws_t_f_gard = wb_template_f['Gard']
ws_t_fund_f_t30tri = wb_template_f['Fundatii_T30tri']


# Fill the modules
antennae_numbers = ws_t_gd['D13'].value + ws_t_gd['D14'].value
module_numbers = ws_t_gd['D17'].value

cell_range = "I25:I32"
start_cell = ws_t_gd[cell_range.split(":")[0]]
end_cell = ws_t_gd[cell_range.split(":")[1]]
current_cell = start_cell

mod_per_ant = module_numbers // antennae_numbers
remainder = module_numbers % antennae_numbers

for i in range(antennae_numbers):
    if i < remainder:
        current_cell.value = mod_per_ant + 1
    else:
        current_cell.value = mod_per_ant
    current_cell = current_cell.offset(1, 0)


# Fill the Formula.xls cells from "Sablon"
# Date generate
ws_t_f_gd['C5'].value = ws_t_gd['C5'].value
ws_t_f_gd['C6'].value = ws_t_gd['C6'].value
ws_t_f_gd['C7'].value = ws_t_gd['C7'].value
ws_t_f_gd['C8'].value = ws_t_gd['C8'].value
ws_t_f_gd['C9'].value = ws_t_gd['C9'].value

ws_t_f_gd['C13'].value = ws_t_gd['C13'].value
ws_t_f_gd['C14'].value = ws_t_gd['C14'].value
ws_t_f_gd['D13'].value = ws_t_gd['D13'].value
ws_t_f_gd['D14'].value = ws_t_gd['D14'].value
ws_t_f_gd['D15'].value = ws_t_gd['D15'].value
ws_t_f_gd['D16'].value = ws_t_gd['D16'].value
ws_t_f_gd['E13'].value = ws_t_gd['E13'].value
ws_t_f_gd['E14'].value = ws_t_gd['E14'].value

ws_t_f_gd['E25'].value = ws_t_gd['E25'].value
ws_t_f_gd['E26'].value = ws_t_gd['E26'].value
ws_t_f_gd['E27'].value = ws_t_gd['E27'].value
ws_t_f_gd['E28'].value = ws_t_gd['E28'].value
ws_t_f_gd['E29'].value = ws_t_gd['E29'].value
ws_t_f_gd['E30'].value = ws_t_gd['E30'].value
ws_t_f_gd['E31'].value = ws_t_gd['E31'].value
ws_t_f_gd['E32'].value = ws_t_gd['E32'].value
ws_t_f_gd['E33'].value = ws_t_gd['E33'].value
ws_t_f_gd['E34'].value = ws_t_gd['E34'].value
ws_t_f_gd['E35'].value = ws_t_gd['E35'].value

ws_t_f_gd['I25'].value = ws_t_gd['I25'].value
ws_t_f_gd['I26'].value = ws_t_gd['I26'].value
ws_t_f_gd['I27'].value = ws_t_gd['I27'].value
ws_t_f_gd['I28'].value = ws_t_gd['I28'].value
ws_t_f_gd['I29'].value = ws_t_gd['I29'].value
ws_t_f_gd['I30'].value = ws_t_gd['I30'].value
ws_t_f_gd['I31'].value = ws_t_gd['I31'].value
ws_t_f_gd['I32'].value = ws_t_gd['I32'].value

ws_t_f_gd['C33'].value = ws_t_gd['C33'].value
ws_t_f_gd['C34'].value = ws_t_gd['C34'].value
ws_t_f_gd['C35'].value = ws_t_gd['C35'].value

ws_t_f_gd['D33'].value = ws_t_gd['D33'].value
ws_t_f_gd['D34'].value = ws_t_gd['D34'].value
ws_t_f_gd['D35'].value = ws_t_gd['D35'].value

ws_t_f_gd['C38'].value = ws_t_gd['C38'].value
ws_t_f_gd['C39'].value = ws_t_gd['C39'].value
ws_t_f_gd['C40'].value = ws_t_gd['C40'].value
ws_t_f_gd['C41'].value = ws_t_gd['C41'].value
ws_t_f_gd['C42'].value = ws_t_gd['C42'].value
ws_t_f_gd['C43'].value = ws_t_gd['C43'].value

ws_t_f_gd['C46'].value = ws_t_gd['C46'].value
ws_t_f_gd['D46'].value = ws_t_gd['D46'].value
ws_t_f_gd['E46'].value = ws_t_gd['E46'].value

ws_t_f_gd['C49'].value = ws_t_gd['C49'].value
ws_t_f_gd['C50'].value = ws_t_gd['C50'].value
ws_t_f_gd['C51'].value = ws_t_gd['C51'].value
ws_t_f_gd['C52'].value = ws_t_gd['C52'].value
ws_t_f_gd['C53'].value = ws_t_gd['C53'].value

ws_t_f_gd['C60'].value = ws_t_gd['C60'].value
ws_t_f_gd['C61'].value = ws_t_gd['C61'].value

ws_t_f_gd['D66'].value = ws_t_gd['D66'].value
ws_t_f_gd['D67'].value = ws_t_gd['D67'].value
ws_t_f_gd['D68'].value = ws_t_gd['D68'].value

ws_t_f_gd['C70'].value = ws_t_gd['C70'].value
ws_t_f_gd['C73'].value = ws_t_gd['C73'].value

ws_t_f_gd['C85'].value = ws_t_gd['C85'].value
ws_t_f_gd['C86'].value = ws_t_gd['C86'].value


# Antene
ws_t_f_a['F2'].value = ws_t_a['F2'].value
ws_t_f_a['G19'].value = ws_t_a['G19'].value
ws_t_f_a['G20'].value = ws_t_a['G20'].value
ws_t_f_a['G28'].value = ws_t_a['G28'].value
ws_t_f_a['G29'].value = ws_t_a['G29'].value
ws_t_f_a['G37'].value = ws_t_a['G37'].value
ws_t_f_a['G38'].value = ws_t_a['G38'].value

ws_t_f_a['M19'].value = ws_t_a['M19'].value
ws_t_f_a['M20'].value = ws_t_a['M20'].value
ws_t_f_a['M28'].value = ws_t_a['M28'].value
ws_t_f_a['M29'].value = ws_t_a['M29'].value
ws_t_f_a['M37'].value = ws_t_a['M37'].value
ws_t_f_a['M38'].value = ws_t_a['M38'].value

# Gard
ws_t_f_gard['G1'].value = ws_t_gard['G1'].value
ws_t_f_gard['J1'].value = ws_t_gard['J1'].value

# Fundatii T30 tri
ws_t_fund_f_t30tri['O10'].value = ws_t_fund_t30tri['O10'].value
ws_t_fund_f_t30tri['P17'].value = ws_t_fund_t30tri['P17'].value
ws_t_fund_f_t30tri['P18'].value = ws_t_fund_t30tri['P18'].value


wb_template_f.save(r'C:\Users\Mihai\Desktop\MyFolder\FormuleSablon.xlsx')
wb_template_f.close()


wb_template_f1 = load_workbook(
    r'C:\Users\Mihai\Desktop\MyFolder\FormuleSablon.xlsx')
ws_t_f_gd1 = wb_template_f1['DateGenerale']


# book = xw.Book(r'C:\Users\Mihai\Desktop\MyFolder\FormuleSablon.xlsx')
xlapp = xw.App(visible=False)
wbxl = xlapp.books.open(r'C:\Users\Mihai\Desktop\MyFolder\FormuleSablon.xlsx')
sheet = wbxl.sheets[0]

row = [sheet.range(3, j).value for j in range(
    1, sheet.api.UsedRange.Columns.Count + 1)]


# Create the T30 file for exchange data to autocad
values = []
name = ''
if ws_t_gd['C5'].value == "T30m tri":
    sheet_T30tri = wbxl.sheets[2]
    values = sheet_T30tri.range("B1:D120").value
    name = "T30tri"
if ws_t_gd['C5'].value == "T20m":
    sheet_T20 = wbxl.sheets[3]
    values = sheet_T20.range("B1:D120").value
    name = "T20"


# Copy Antennae Table
sheet1 = wbxl.sheets[1]
values2 = sheet1.range("E5:L15").value
for i, row in enumerate(values2):
    for j, val in enumerate(row):
        ws_t_at.cell(row=i+2, column=j+1, value=val)


wbxl.close()
xlapp.quit()

#  Load the Excel file into a Pandas DataFrame
df = pd.DataFrame(values)
df = df.dropna()

#  Write the values to a text file
df.to_csv(r'C:\Users\Mihai\Desktop\MyFolder\{}.txt'.format(
    name), sep="\t", index=False, header=None)

# In Date Generale add the row 3 data for word template
# row_3_f_gd = []
# for cell in ws_t_f_gd1[3]:
#     row_3_f_gd.append(cell.value)

# for i, cell in enumerate(ws_t_gd[3]):
#     cell.value = row[i]


wb_template.save(r'C:\Users\Mihai\Desktop\MyFolder\Sablon.xlsx')


wb_template_f1.close()
wb_template.close()
